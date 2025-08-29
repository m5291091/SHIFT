from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib.auth.models import Group
from .forms import SignUpForm

from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from datetime import date, datetime, time, timedelta
from collections import defaultdict

from .models import Member, Assignment, LeaveRequest, MemberAvailability, ShiftPattern, OtherAssignment, TimeSlotRequirement, FixedAssignment, Department, DesignatedHoliday, SolverSettings, PaidLeave
from .serializers import MemberSerializer, AssignmentSerializer, MemberAvailabilitySerializer, ShiftPatternSerializer, OtherAssignmentSerializer, FixedAssignmentSerializer, DepartmentSerializer, DesignatedHolidaySerializer, SolverSettingsSerializer, PaidLeaveSerializer
from .solver import generate_schedule

def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = True
            user.save()

            # Add user to the "Default Staff" group
            try:
                staff_group = Group.objects.get(name='Default Staff')
                user.groups.add(staff_group)
            except Group.DoesNotExist:
                # This case should ideally not happen if migrations are run correctly
                pass

            login(request, user)
            return redirect('admin:index')
    else:
        form = SignUpForm()
    return render(request, 'registration/signup.html', {'form': form})

# Base class for user-filtered data
class UserFilteredListView(generics.ListAPIView):
    def get_queryset(self):
        """Filter queryset by the logged-in user."""
        return self.queryset.filter(created_by=self.request.user)

    def perform_create(self, serializer):
        """Associate the object with the logged-in user upon creation."""
        serializer.save(created_by=self.request.user)

class DepartmentListView(UserFilteredListView):
    queryset = Department.objects.all().order_by('name')
    serializer_class = DepartmentSerializer

class MemberListView(UserFilteredListView):
    queryset = Member.objects.all().order_by('sort_order', 'name')
    serializer_class = MemberSerializer

class ShiftPatternListView(UserFilteredListView):
    serializer_class = ShiftPatternSerializer
    queryset = ShiftPattern.objects.all()

    def get_queryset(self):
        """Additionally filter by department if provided."""
        queryset = super().get_queryset() # Apply user filter first
        department_id = self.request.query_params.get('department_id')
        if department_id is not None:
            queryset = queryset.filter(department_id=department_id)
        return queryset

class ScheduleDataView(APIView):
    def get(self, request, *args, **kwargs):
        department_id = self.request.query_params.get('department_id')
        start_date_str = self.request.query_params.get('start_date')
        end_date_str = self.request.query_params.get('end_date')
        
        # Filter by the logged-in user
        user = self.request.user
        members_queryset = Member.objects.filter(created_by=user)
        shift_patterns_queryset = ShiftPattern.objects.filter(created_by=user)

        if department_id:
            members_queryset = members_queryset.filter(department_id=department_id)
            shift_patterns_queryset = shift_patterns_queryset.filter(department_id=department_id)

        members = members_queryset.order_by('sort_order', 'name')
        member_serializer = MemberSerializer(members, many=True)

        if not start_date_str or not end_date_str:
            return Response({
                'members': member_serializer.data,
                'assignments': [], 'leave_requests': [], 'availabilities': [],
                'other_assignments': [], 'earnings': {},
                'fixed_assignments': [], 'designated_holidays': [], 'paid_leaves': [],
            })

        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)

        # Filter all data by the logged-in user
        assignments = Assignment.objects.filter(
            shift_date__range=[start_date, end_date],
            member__department_id=department_id,
            created_by=user
        ).select_related('member', 'shift_pattern')
        assignment_serializer = AssignmentSerializer(assignments, many=True)

        fixed_assignments = FixedAssignment.objects.filter(
            shift_date__range=[start_date, end_date],
            member__department_id=department_id,
            created_by=user
        ).select_related('member', 'shift_pattern')
        fixed_assignment_serializer = FixedAssignmentSerializer(fixed_assignments, many=True)

        leave_requests = LeaveRequest.objects.filter(
            status='approved', 
            leave_date__range=[start_date, end_date],
            member__department_id=department_id,
            created_by=user
        )
        leave_data = [{'leave_date': str(req.leave_date), 'member_id': req.member.id} for req in leave_requests]
        
        availabilities = MemberAvailability.objects.filter(
            member__department_id=department_id,
            created_by=user
        )
        availability_serializer = MemberAvailabilitySerializer(availabilities, many=True)

        other_assignments = OtherAssignment.objects.filter(
            shift_date__range=[start_date, end_date],
            member__department_id=department_id,
            created_by=user
        )
        other_serializer = OtherAssignmentSerializer(other_assignments, many=True)

        designated_holidays = DesignatedHoliday.objects.filter(
            date__range=[start_date, end_date],
            member__department_id=department_id,
            created_by=user
        )
        designated_holiday_serializer = DesignatedHolidaySerializer(designated_holidays, many=True)

        paid_leaves = PaidLeave.objects.filter(
            date__range=[start_date, end_date],
            member__department_id=department_id,
            created_by=user
        ).select_related('member')
        paid_leave_serializer = PaidLeaveSerializer(paid_leaves, many=True)

        earnings_map = defaultdict(float)
        premium_start, premium_end = time(22, 0), time(5, 0)
        for assign in assignments:
            if assign.member.employee_type == 'hourly' and assign.member.hourly_wage:
                p = assign.shift_pattern
                total_duration = (datetime.combine(date.today(), p.end_time) - datetime.combine(date.today(), p.start_time)).total_seconds() / 60
                if p.end_time < p.start_time: total_duration += 24 * 60
                work_minutes = total_duration - p.break_minutes
                
                premium_minutes = 0
                current_time = datetime.combine(date.today(), p.start_time)
                for _ in range(int(total_duration)):
                    if current_time.time() >= premium_start or current_time.time() < premium_end:
                        premium_minutes += 1
                    current_time += timedelta(minutes=1)
                
                normal_minutes = work_minutes - premium_minutes
                
                earnings = (normal_minutes * assign.member.hourly_wage) + (premium_minutes * assign.member.hourly_wage * 1.25)
                earnings_map[assign.member.id] += earnings / 60
        
        for pl in paid_leaves:
            if pl.member.employee_type == 'hourly' and pl.member.hourly_wage:
                earnings_map[pl.member.id] += (pl.hours * pl.member.hourly_wage)
        
        return Response({
            'assignments': assignment_serializer.data,
            'fixed_assignments': fixed_assignment_serializer.data,
            'leave_requests': leave_data,
            'members': member_serializer.data,
            'availabilities': availability_serializer.data,
            'other_assignments': other_serializer.data,
            'designated_holidays': designated_holiday_serializer.data,
            'paid_leaves': paid_leave_serializer.data,
            'earnings': {k: round(v) for k, v in earnings_map.items()},
        })

class GenerateShiftView(APIView):
    def post(self, request, *args, **kwargs):
        department_id = request.data.get('department_id')
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')
        if not start_date_str or not end_date_str or not department_id:
            return Response({'error': 'department_id, start_date and end_date are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not Department.objects.filter(id=department_id, created_by=request.user).exists():
            return Response({'error': 'Invalid department'}, status=status.HTTP_403_FORBIDDEN)

        # Run the solver
        result = generate_schedule(department_id, start_date_str, end_date_str)

        if result.get('success'):
            # Clear previous assignments for this period
            start_date = date.fromisoformat(start_date_str)
            end_date = date.fromisoformat(end_date_str)
            Assignment.objects.filter(
                shift_date__range=[start_date, end_date],
                member__department_id=department_id,
                created_by=request.user
            ).delete()

            # Create new assignments from solver result, adding the current user
            assignments_data = result.get('assignments', [])
            new_assignments = []
            for assign_data in assignments_data:
                new_assignments.append(
                    Assignment(
                        member_id=assign_data['member_id'],
                        shift_pattern_id=assign_data['shift_pattern_id'],
                        shift_date=assign_data['shift_date'],
                        created_by=request.user
                    )
                )
            
            if new_assignments:
                Assignment.objects.bulk_create(new_assignments)
            
            # Serialize the newly created objects to return to the frontend
            serializer = AssignmentSerializer(new_assignments, many=True)
            response_data = {
                'success': True,
                'infeasible_days': result.get('infeasible_days', {}),
                'assignments': serializer.data
            }
            return Response(response_data, status=status.HTTP_200_OK)
        else:
            # If solver failed, just return the failure message
            return Response(result, status=status.HTTP_200_OK)

class ManualAssignmentView(APIView):
    def post(self, request, *args, **kwargs):
        member_id = request.data.get('member_id')
        shift_date = request.data.get('shift_date')
        pattern_id = request.data.get('pattern_id')
        
        member = Member.objects.get(id=member_id)
        if member.created_by != request.user:
            return Response(status=status.HTTP_403_FORBIDDEN)

        DesignatedHoliday.objects.filter(member_id=member_id, date=shift_date).delete()
        PaidLeave.objects.filter(member_id=member_id, date=shift_date).delete()
        
        Assignment.objects.filter(member_id=member_id, shift_date=shift_date).delete()
        OtherAssignment.objects.filter(member_id=member_id, shift_date=shift_date).delete()
        if pattern_id:
            Assignment.objects.create(member_id=member_id, shift_pattern_id=pattern_id, shift_date=shift_date, created_by=request.user)
        return Response(status=status.HTTP_200_OK)

class OtherAssignmentView(APIView):
    def post(self, request, *args, **kwargs):
        member_id = request.data.get('member_id')
        shift_date = request.data.get('shift_date')
        activity_name = request.data.get('activity_name')
        
        member = Member.objects.get(id=member_id)
        if member.created_by != request.user:
            return Response(status=status.HTTP_403_FORBIDDEN)

        DesignatedHoliday.objects.filter(member_id=member_id, date=shift_date).delete()
        PaidLeave.objects.filter(member_id=member_id, date=shift_date).delete()
        
        Assignment.objects.filter(member_id=member_id, shift_date=shift_date).delete()
        OtherAssignment.objects.filter(member_id=member_id, shift_date=shift_date).delete()
        if activity_name:
            OtherAssignment.objects.create(member_id=member_id, shift_date=shift_date, activity_name=activity_name, created_by=request.user)
        return Response(status=status.HTTP_200_OK)

class BulkFixedAssignmentView(APIView):
    def post(self, request, *args, **kwargs):
        assignments = request.data.get('assignments', [])
        if not assignments:
            return Response({'error': 'No assignments provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Security check: ensure all members belong to the current user
        member_ids = {assign['member_id'] for assign in assignments}
        if Member.objects.filter(id__in=member_ids, created_by=request.user).count() != len(member_ids):
            return Response({'error': 'Invalid member ID included'}, status=status.HTTP_403_FORBIDDEN)

        fixed_assignments_to_create = []
        for assign in assignments:
            fixed_assignments_to_create.append(
                FixedAssignment(
                    member_id=assign['member_id'],
                    shift_pattern_id=assign['shift_pattern_id'],
                    shift_date=assign['shift_date'],
                    created_by=request.user
                )
            )
        
        FixedAssignment.objects.bulk_create(fixed_assignments_to_create, ignore_conflicts=True)
        
        return Response({'status': 'success'}, status=status.HTTP_201_CREATED)

class FixedAssignmentView(APIView):
    def post(self, request, *args, **kwargs):
        member_id = request.data.get('member_id')
        shift_date = request.data.get('shift_date')
        pattern_id = request.data.get('pattern_id')
        
        if not all([member_id, shift_date]):
            return Response({'error': 'member_id and shift_date are required'}, status=status.HTTP_400_BAD_REQUEST)

        member = Member.objects.get(id=member_id)
        if member.created_by != request.user:
            return Response(status=status.HTTP_403_FORBIDDEN)

        DesignatedHoliday.objects.filter(member_id=member_id, date=shift_date).delete()
        PaidLeave.objects.filter(member_id=member_id, date=shift_date).delete()
        
        Assignment.objects.filter(member_id=member_id, shift_date=shift_date).delete()
        
        if pattern_id:
            FixedAssignment.objects.update_or_create(
                member_id=member_id,
                shift_date=shift_date,
                defaults={'shift_pattern_id': pattern_id, 'created_by': request.user}
            )
        else:
            FixedAssignment.objects.filter(member_id=member_id, shift_date=shift_date).delete()
            
        return Response(status=status.HTTP_200_OK)

class DesignatedHolidayView(APIView):
    def post(self, request, *args, **kwargs):
        member_id = request.data.get('member_id')
        date = request.data.get('date')

        if not all([member_id, date]):
            return Response({'error': 'member_id and date are required'}, status=status.HTTP_400_BAD_REQUEST)

        member = Member.objects.get(id=member_id)
        if member.created_by != request.user:
            return Response(status=status.HTTP_403_FORBIDDEN)

        holiday, created = DesignatedHoliday.objects.get_or_create(
            member_id=member_id,
            date=date,
            defaults={'created_by': request.user}
        )

        if created:
            Assignment.objects.filter(member_id=member_id, shift_date=date).delete()
            FixedAssignment.objects.filter(member_id=member_id, shift_date=date).delete()
            OtherAssignment.objects.filter(member_id=member_id, shift_date=date).delete()
            PaidLeave.objects.filter(member_id=member_id, date=date).delete()
            action = "created"
        else:
            action = "already_exists"
            
        return Response({'status': f'holiday {action}'}, status=status.HTTP_200_OK)

class PaidLeaveView(APIView):
    def post(self, request, *args, **kwargs):
        member_id = request.data.get('member_id')
        date = request.data.get('date')

        if not all([member_id, date]):
            return Response({'error': 'member_id and date are required'}, status=status.HTTP_400_BAD_REQUEST)

        member = Member.objects.get(id=member_id)
        if member.created_by != request.user:
            return Response(status=status.HTTP_403_FORBIDDEN)

        Assignment.objects.filter(member_id=member_id, shift_date=date).delete()
        FixedAssignment.objects.filter(member_id=member_id, shift_date=date).delete()
        OtherAssignment.objects.filter(member_id=member_id, shift_date=date).delete()
        DesignatedHoliday.objects.filter(member_id=member_id, date=date).delete()
        LeaveRequest.objects.filter(member_id=member_id, leave_date=date).delete()

        paid_leave, created = PaidLeave.objects.get_or_create(
            member_id=member_id,
            date=date,
            defaults={'created_by': request.user}
        )

        if created:
            action = "created"
        else:
            action = "already_exists"
            
        return Response({'status': f'paid_leave {action}'}, status=status.HTTP_200_OK)

    def delete(self, request, *args, **kwargs):
        member_id = request.data.get('member_id')
        date = request.data.get('date')

        if not all([member_id, date]):
            return Response({'error': 'member_id and date are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        deleted_count, _ = PaidLeave.objects.filter(member_id=member_id, date=date, created_by=request.user).delete()
        if deleted_count > 0:
            return Response({'status': 'deleted'}, status=status.HTTP_200_OK)
        return Response({'status': 'not_found'}, status=status.HTTP_404_NOT_FOUND)

class SolverSettingsDetailView(generics.RetrieveUpdateAPIView):
    queryset = SolverSettings.objects.all()
    serializer_class = SolverSettingsSerializer
    lookup_field = 'department_id'

    def get_queryset(self):
        return self.queryset.filter(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(created_by=self.request.user)

class SolverSettingsListView(generics.ListCreateAPIView):
    queryset = SolverSettings.objects.all()
    serializer_class = SolverSettingsSerializer

    def get_queryset(self):
        return self.queryset.filter(created_by=self.request.user)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class BulkAssignmentDeleteView(APIView):
    def post(self, request, *args, **kwargs):
        assignment_ids = request.data.get('assignment_ids', [])
        if not assignment_ids:
            return Response({'error': 'No assignment IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        Assignment.objects.filter(id__in=assignment_ids, created_by=request.user).delete()
        
        return Response({'status': 'success'}, status=status.HTTP_200_OK)

class BulkFixedAssignmentDeleteView(APIView):
    def post(self, request, *args, **kwargs):
        fixed_assignment_ids = request.data.get('fixed_assignment_ids', [])
        if not fixed_assignment_ids:
            return Response({'error': 'No fixed assignment IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        FixedAssignment.objects.filter(id__in=fixed_assignment_ids, created_by=request.user).delete()
        
        return Response({'status': 'success'}, status=status.HTTP_200_OK)

class BulkOtherAssignmentDeleteView(APIView):
    def post(self, request, *args, **kwargs):
        other_assignment_ids = request.data.get('other_assignment_ids', [])
        if not other_assignment_ids:
            return Response({'error': 'No other assignment IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        OtherAssignment.objects.filter(id__in=other_assignment_ids, created_by=request.user).delete()
        
        return Response({'status': 'success'}, status=status.HTTP_200_OK)

class BulkDesignatedHolidayDeleteView(APIView):
    def post(self, request, *args, **kwargs):
        designated_holiday_ids = request.data.get('designated_holiday_ids', [])
        if not designated_holiday_ids:
            return Response({'error': 'No designated holiday IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        DesignatedHoliday.objects.filter(id__in=designated_holiday_ids, created_by=request.user).delete()
        
        return Response({'status': 'success'}, status=status.HTTP_200_OK)

class BulkPaidLeaveDeleteView(APIView):
    def post(self, request, *args, **kwargs):
        paid_leave_ids = request.data.get('paid_leave_ids', [])
        if not paid_leave_ids:
            return Response({'error': 'No paid leave IDs provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        PaidLeave.objects.filter(id__in=paid_leave_ids, created_by=request.user).delete()
        
        return Response({'status': 'success'}, status=status.HTTP_200_OK)

from django.http import HttpResponse
import openpyxl
from openpyxl import utils
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from io import BytesIO

class ShiftExportExcelView(APIView):
    def get(self, request, *args, **kwargs):
        department_id = self.request.query_params.get('department_id')
        start_date_str = self.request.query_params.get('start_date')
        end_date_str = self.request.query_params.get('end_date')
        
        if not all([department_id, start_date_str, end_date_str]):
            return Response({'error': 'department_id, start_date, and end_date are required'}, status=status.HTTP_400_BAD_REQUEST)

        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        user = self.request.user

        # Fetch data
        members = Member.objects.filter(created_by=user, department_id=department_id).order_by('sort_order', 'name')
        assignments = Assignment.objects.filter(
            created_by=user, member__department_id=department_id, shift_date__range=[start_date, end_date]
        ).select_related('member', 'shift_pattern')
        other_assignments = OtherAssignment.objects.filter(
            created_by=user, member__department_id=department_id, shift_date__range=[start_date, end_date]
        ).select_related('member')
        paid_leaves = PaidLeave.objects.filter(
            created_by=user, member__department_id=department_id, date__range=[start_date, end_date]
        ).select_related('member')
        designated_holidays = DesignatedHoliday.objects.filter(
            created_by=user, member__department_id=department_id, date__range=[start_date, end_date]
        ).select_related('member')

        # Data processing
        shift_data = defaultdict(dict)
        for assignment in assignments:
            if assignment.shift_pattern:
                shift_data[assignment.shift_date][assignment.member_id] = assignment.shift_pattern.short_name
        for other in other_assignments:
            shift_data[other.shift_date][other.member_id] = other.activity_name
        for leave in paid_leaves:
            shift_data[leave.date][leave.member_id] = "有給"
        for holiday in designated_holidays:
            shift_data[holiday.date][holiday.member_id] = "公休"

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{start_date.strftime('%Y年%m月')} シフト表"

        # Header
        header = ['日付', '曜日'] + [member.name for member in members]
        ws.append(header)

        # Styles
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Body
        saturday_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        sunday_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        
        current_date = start_date
        while current_date <= end_date:
            day_of_week = current_date.strftime('%a')
            row_data = [current_date.strftime('%d'), day_of_week]
            
            for member in members:
                row_data.append(shift_data[current_date].get(member.id, ''))

            ws.append(row_data)
            
            # Apply styles to the row
            row_index = ws.max_row
            for cell in ws[row_index]:
                cell.alignment = center_align
                cell.border = thin_border
            
            if current_date.weekday() == 5: # Saturday
                for cell in ws[row_index]:
                    cell.fill = saturday_fill
            elif current_date.weekday() == 6: # Sunday
                for cell in ws[row_index]:
                    cell.fill = sunday_fill

            current_date += timedelta(days=1)

        # Adjust column widths
        for i, column_cells in enumerate(ws.columns):
            max_length = 0
            column = openpyxl.utils.get_column_letter(i + 1)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="shift_schedule_{start_date.strftime("%Y%m")}.xlsx"'
        
        return response
